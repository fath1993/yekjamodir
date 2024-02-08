import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.contrib.sites.models import Site
import io
import jdatetime
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.shortcuts import render, redirect
from django.views import View
from openpyxl.utils import get_column_letter

from financial_accounting.models import TransactionRecord, FinancialBroker
from gallery.models import FileGallery
from gallery.views import file_processor, file_allowed_to_upload, user_quote_limit_exceed
from utilities.http_metod import fetch_data_from_http_post, fetch_multiple_files_from_http_file, \
    fetch_datalist_from_http_post


class FinancialBrokerNew(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'ثبت کارگزار جدید',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_main_item_id': 'financial',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'ثبت کارگزار جدید',
                        }

    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return render(request, 'financial/financial-broker-new.html', self.context)
        else:
            return redirect('accounts:login')

    def post(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            broker_name = fetch_data_from_http_post(request, 'broker-name', self.context)
            account_owner = fetch_data_from_http_post(request, 'account-owner', self.context)
            account_number = fetch_data_from_http_post(request, 'account-number', self.context)
            isbn = fetch_data_from_http_post(request, 'isbn', self.context)
            account_card_number = fetch_data_from_http_post(request, 'account-card-number', self.context)
            if broker_name is None:
                self.context['alert'] = 'نام کارگزار بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            if account_owner is None:
                self.context['alert'] = 'نام صاحب حساب بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            if account_number is None:
                self.context['alert'] = 'شناسه(شماره) حساب بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            try:
                broker = FinancialBroker.objects.get(broker_name=broker_name, account_owner=account_owner,
                                                     account_number=account_number, created_by=request.user)
                self.context['alert'] = 'حساب با مشخصات فوق از قبل ثبت شده است'
                self.context['toast_danger'] = 'حساب با مشخصات فوق از قبل ثبت شده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            except Exception as e:
                print(str(e))
                new_broker = FinancialBroker(
                    broker_name=broker_name,
                    account_owner=account_owner,
                    account_number=account_number,
                    account_ISBN=isbn,
                    account_card_number=account_card_number,
                    created_by=request.user,
                    updated_by=request.user,
                )
                new_broker.save()
            return redirect('financial:financial-broker-list')
        else:
            return redirect('accounts:login')


class FinancialBrokerEdit(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'ویرایش کارگزار',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_main_item_id': 'financial',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'ویرایش کارگزار',
                        }

    def get(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                self.context['broker'] = broker
                return render(request, 'financial/financial-broker-edit.html', self.context)
            except:
                return render(request, '404.html')
        else:
            return redirect('accounts:login')

    def post(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                self.context['broker'] = broker
            except:
                return render(request, '404.html')
            broker_name = fetch_data_from_http_post(request, 'broker-name', self.context)
            account_owner = fetch_data_from_http_post(request, 'account-owner', self.context)
            account_number = fetch_data_from_http_post(request, 'account-number', self.context)
            isbn = fetch_data_from_http_post(request, 'isbn', self.context)
            account_card_number = fetch_data_from_http_post(request, 'account-card-number', self.context)
            if broker_name is None:
                self.context['alert'] = 'نام کارگزار بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            if account_owner is None:
                self.context['alert'] = 'نام صاحب حساب بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            if account_number is None:
                self.context['alert'] = 'شناسه(شماره) حساب بدرستی وارد نشده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            try:
                broker = FinancialBroker.objects.get(broker_name=broker_name, account_owner=account_owner,
                                                     account_number=account_number, created_by=request.user)
                self.context['alert'] = 'حساب با مشخصات فوق از قبل ثبت شده است'
                self.context['toast_danger'] = 'حساب با مشخصات فوق از قبل ثبت شده است'
                return render(request, 'financial/financial-broker-new.html', self.context)
            except Exception as e:
                broker.broker_name = broker_name
                broker.account_owner = account_owner
                broker.account_number = account_number
                broker.account_ISBN = isbn
                broker.account_card_number = account_card_number
                broker.save()
            return redirect('financial:financial-broker-edit-with-broker-id', broker_id=broker_id)
        else:
            return redirect('accounts:login')


class FinancialBrokerList(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'لیست کارگزار های مالی',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_sub_item_id': 'financial-broker-list',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'لیست کارگزار های مالی',
                        }

    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            print(brokers)
            return render(request, 'financial/financial-brokers-list.html', self.context)
        else:
            return redirect('accounts:login')

    def post(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return render(request, '404.html')
        else:
            return redirect('accounts:login')


class FinancialBrokerRemove(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def get(self, request, broker_id, *args, **kwargs):
        return render(request, '404.html')

    def post(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            if request.method == 'POST':
                try:
                    broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                    broker.delete()
                except Exception as e:
                    print(str(e))
                    return HttpResponse(f'delete error: {str(e)}')
            else:
                pass
            return redirect('financial:financial-broker-list')
        else:
            return redirect('accounts:login')


class FinancialTransactionRecordNew(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'ثبت فاکتور جدید',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_sub_item_id': 'financial-transaction-record-new',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'ثبت فاکتور جدید',
                        }

    def get(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                self.context['broker'] = broker
            except:
                return render(request, '404.html')
            return render(request, 'financial/financial-transaction-record-new.html', self.context)
        else:
            return redirect('accounts:login')

    def post(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                self.context['broker'] = broker
            except:
                return render(request, '404.html')

            transaction_type = fetch_data_from_http_post(request, 'transaction_type', self.context)
            transaction_title = fetch_data_from_http_post(request, 'transaction_title', self.context)
            transaction_description = fetch_data_from_http_post(request, 'transaction_description', self.context)
            transaction_date = fetch_data_from_http_post(request, 'transaction_date', self.context)
            transaction_amount = fetch_data_from_http_post(request, 'transaction_amount', self.context)
            uploaded_image_ids = fetch_datalist_from_http_post(request, 'uploaded_image_ids', self.context)

            if transaction_type is None:
                self.context['alert'] = 'نوع فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-new.html', self.context)
            if transaction_title is None:
                self.context['alert'] = 'عنوان فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-new.html', self.context)
            if transaction_date is None:
                self.context['alert'] = 'تاریخ فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-new.html', self.context)
            else:
                transaction_date = str(transaction_date).split('/')
                transaction_date = jdatetime.datetime(year=int(transaction_date[0]), month=int(transaction_date[1]),
                                                      day=int(transaction_date[2]))
            if transaction_amount is None:
                self.context['alert'] = 'مقدار ریالی فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-new.html', self.context)
            else:
                try:
                    transaction_amount = str(transaction_amount).replace(',', '')
                    transaction_amount = int(transaction_amount)
                except:
                    self.context['alert'] = 'مقدار ریالی فاکتور بدرستی وارد نشده است'
                    return render(request, 'financial/financial-transaction-record-new.html', self.context)

            new_transaction_record = TransactionRecord(
                financial_broker=broker,
                transaction_type=transaction_type,
                amount=transaction_amount,
                title=transaction_title,
                date_of_action=transaction_date,
                description=transaction_description,
                created_by=request.user,
                updated_by=request.user,
            )
            new_transaction_record.save()

            if uploaded_image_ids:
                for file_id in uploaded_image_ids:
                    file = FileGallery.objects.get(id=file_id)
                    new_transaction_record.attachments.add(file)
                    new_transaction_record.save()

            transaction_records = TransactionRecord.objects.filter(financial_broker=broker)
            self.context['transaction_records'] = transaction_records

            return redirect('financial:financial-transaction-record-list-with-broker-id', broker_id=broker_id)
        else:
            return redirect('accounts:login')


class FinancialTransactionRecordEdit(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'ویرایش صورت حساب',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_main_item_id': 'financial',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'ویرایش صورت حساب',
                        }

    def get(self, request, transaction_id, *args, **kwargs):
        if request.user.is_authenticated:
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            try:
                transaction = TransactionRecord.objects.get(id=transaction_id, created_by=request.user)
                self.context['transaction'] = transaction
                self.context['broker'] = transaction.financial_broker
                return render(request, 'financial/financial-transaction-record-edit.html', self.context)
            except Exception as e:
                print(str(e))
                return render(request, '404.html')
        else:
            return redirect('accounts:login')

    def post(self, request, transaction_id, *args, **kwargs):
        if request.user.is_authenticated:
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            try:
                transaction = TransactionRecord.objects.get(id=transaction_id, created_by=request.user)
                self.context['transaction'] = transaction
                broker = transaction.financial_broker
                self.context['broker'] = broker
            except:
                return render(request, '404.html')
            transaction_type = fetch_data_from_http_post(request, 'transaction_type', self.context)
            transaction_title = fetch_data_from_http_post(request, 'transaction_title', self.context)
            transaction_description = fetch_data_from_http_post(request, 'transaction_description', self.context)
            transaction_date = fetch_data_from_http_post(request, 'transaction_date', self.context)
            transaction_amount = fetch_data_from_http_post(request, 'transaction_amount', self.context)
            uploaded_image_ids = fetch_datalist_from_http_post(request, 'uploaded_image_ids', self.context)

            if transaction_type is None:
                self.context['alert'] = 'نوع فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-edit.html', self.context)
            if transaction_title is None:
                self.context['alert'] = 'عنوان فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-edit.html', self.context)
            if transaction_date is None:
                self.context['alert'] = 'تاریخ فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-edit.html', self.context)
            else:
                transaction_date = str(transaction_date).split('/')
                transaction_date = jdatetime.datetime(year=int(transaction_date[0]), month=int(transaction_date[1]),
                                                      day=int(transaction_date[2]))
            if transaction_amount is None:
                self.context['alert'] = 'مقدار ریالی فاکتور بدرستی وارد نشده است'
                return render(request, 'financial/financial-transaction-record-edit.html', self.context)
            else:
                try:
                    transaction_amount = str(transaction_amount).replace(',', '')
                    transaction_amount = int(transaction_amount)
                except:
                    self.context['alert'] = 'مقدار ریالی فاکتور بدرستی وارد نشده است'
                    return render(request, 'financial/financial-transaction-record-edit.html', self.context)

            transaction.transaction_type = transaction_type
            transaction.amount = transaction_amount
            transaction.title = transaction_title
            transaction.date_of_action = transaction_date
            transaction.description = transaction_description
            transaction.save()

            if uploaded_image_ids:
                for file_id in uploaded_image_ids:
                    file = FileGallery.objects.get(id=file_id)
                    transaction.attachments.add(file)
                    transaction.save()

            return redirect('financial:financial-transaction-record-edit-with-transaction-id',
                            transaction_id=transaction_id)
        else:
            return redirect('accounts:login')


class FinancialTransactionRecordList(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {'page_title': 'لیست صورت حساب ها',
                        'navigation_icon_menu_id': 'service',
                        'navigation_menu_body_id': 'navigationService',
                        'navigation_menu_body_sub_item_id': 'financial-transaction-record-list',
                        'breadcrumb_1': 'خانه',
                        'breadcrumb_2': 'حسابداری',
                        'breadcrumb_3': 'لیست صورت حساب ها',
                        }

    def get(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            profile = request.user.profile_user
            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            if brokers.count() == 0:
                return redirect('financial:financial-broker-new')
            if broker_id == 0:
                return redirect('financial:financial-transaction-record-list-with-broker-id', broker_id=brokers[0].id)
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                transaction_records = TransactionRecord.objects.filter(financial_broker__id=broker_id,
                                                                       created_by=request.user)
                self.context['broker'] = broker
                self.context['transaction_records'] = transaction_records
            except Exception as e:
                print(str(e))
                return render(request, '404.html')

            financial_account_summary = get_financial_account_summery(transaction_records)
            self.context['financial_account_summary'] = financial_account_summary
            return render(request, 'financial/financial-transaction-record-list.html', self.context)
        else:
            return redirect('accounts:login')

    def post(self, request, broker_id, *args, **kwargs):
        if request.user.is_authenticated:
            date_from = fetch_data_from_http_post(request, 'date-from', self.context)
            date_to = fetch_data_from_http_post(request, 'date-to', self.context)

            if date_from:
                date_from = str(date_from).split('/')
                date_from = jdatetime.datetime(year=int(date_from[0]), month=int(date_from[1]), day=int(date_from[2]))
            if date_to:
                date_to = str(date_to).split('/')
                date_to = jdatetime.datetime(year=int(date_to[0]), month=int(date_to[1]), day=int(date_to[2]))

            brokers = FinancialBroker.objects.filter(created_by=request.user)
            self.context['brokers'] = brokers
            try:
                broker = FinancialBroker.objects.get(id=broker_id, created_by=request.user)
                transaction_records = TransactionRecord.objects.filter(financial_broker__id=broker_id,
                                                                       created_by=request.user)
                self.context['broker'] = broker
                self.context['transaction_records'] = transaction_records
            except Exception as e:
                print(str(e))
                return render(request, '404.html')

            filtered_transaction_records = transaction_records.filter(financial_broker__id=broker_id
                                                                      , date_of_action__range=[date_from, date_to])
            self.context['transaction_records'] = filtered_transaction_records
            financial_account_summary = get_financial_account_summery(filtered_transaction_records)
            self.context['financial_account_summary'] = financial_account_summary
            return render(request, 'financial/financial-transaction-record-list.html', self.context)
        else:
            return redirect('accounts:login')


class FinancialTransactionRecordRemove(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def get(self, request, transaction_id, *args, **kwargs):
        return render(request, '404.html')

    def post(self, request, transaction_id, *args, **kwargs):
        if request.user.is_authenticated:
            if request.method == 'POST':
                try:
                    transaction_record = TransactionRecord.objects.get(id=transaction_id, created_by=request.user)
                    for attachment in transaction_record.attachments.all():
                        attachment.delete()
                    transaction_record.delete()
                    return redirect('financial:financial-transaction-record-list-with-broker-id',
                                    broker_id=transaction_record.financial_broker.id)
                except Exception as e:
                    print(str(e))
                    return HttpResponse(f'delete error: {str(e)}')
            else:
                return render(request, '404.html')
        else:
            return redirect('accounts:login')


def get_financial_account_summery(transactions):
    withdraw = 0
    deposit = 0
    total = 0
    state = ''
    for transaction_record in transactions:
        if transaction_record.transaction_type == 'withdraw':
            withdraw += abs(transaction_record.amount) / 10
        else:
            deposit += abs(transaction_record.amount) / 10
    total = abs(withdraw - deposit)
    if deposit > withdraw:
        state = 'green'
    elif deposit < withdraw:
        state = 'red'
    else:
        state = 'zero'
    return f'{int(round(withdraw)):,}', f'{int(round(deposit)):,}', f'{int(round(total)):,}', state


def ajax_export_transaction_to_excel(request):
    if request.user.is_authenticated:
        output = io.BytesIO()
        transactions = request.POST['transactions']
        transactions = str(transactions).split(',')
        del transactions[-1]
        transactions = TransactionRecord.objects.filter(id__in=transactions)
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.append(['تاریخ گزارش', jdatetime.datetime.now().strftime('%Y-%m-%d %H:%M')])

        withdraw = 0
        deposit = 0
        for obj in transactions:
            if obj.transaction_type == 'deposit':
                deposit += int(obj.amount)
            else:
                withdraw += int(obj.amount)
        balance = deposit - withdraw
        if balance < 0:
            balance = 'منفی ' + str(-1 * balance) + ' ریال'
        else:
            balance = 'مثبت  ' + str(balance) + ' ریال'

        ws.append(['برداشت شده: ' + str(withdraw) + ' ریال', 'واریز شده: ' + str(deposit) + ' ریال',
                   'مجموع محاسبه شده: ' + balance])

        ws.append(['اکانت مالی', 'نوع تراکنش', 'مقدار - ریال', 'عنوان', 'تاریخ اقدام',
                   'توضیحات', 'ایجاد شده در', 'بروز شده در', 'ایجاد شده توسط', 'بروز شده توسط', 'ضمائم'])
        i = 3
        for obj in transactions:
            attachments_links = ''
            attachments = obj.attachments.all()
            domain = Site.objects.get_current().domain
            for attachment in attachments:
                attachments_links += f'https://{domain}' + attachment.file.url + '; ' + '\n'

            if obj.transaction_type == 'withdraw':
                transaction_type = 'برداشت از حساب'
            else:
                transaction_type = 'واریز به حساب'
            ws.append(
                [obj.financial_broker.broker_name, transaction_type, obj.amount, obj.title,
                 obj.date_of_action.strftime('%Y-%m-%d %H:%M'), obj.description,
                 obj.created_at.strftime('%Y-%m-%d %H:%M'),
                 obj.updated_at.strftime('%Y-%m-%d %H:%M'), obj.created_by.username, obj.updated_by.username,
                 attachments_links])
            i += 1

        # Set alignment to middle center for all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in range(3, i):
            cell = ws.cell(row=row, column=8)
            cell.alignment = Alignment(wrap_text=True, shrinkToFit=True, justifyLastLine=True)

            # Automatically adjust column widths to fit content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)  # Get the column letter (e.g., 'A', 'B', etc.)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2  # Add a little padding and scaling factor
                ws.column_dimensions[column_letter].width = adjusted_width

        now = jdatetime.datetime.now().strftime('%Y-%m-%d')

        # Save the workbook to the in-memory byte stream
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={now}-transaction-records.xlsx'

        return response


def ajax_set_default_broker(request):
    context = {}
    if request.user.is_authenticated:
        default_broker_id = fetch_data_from_http_post(request, 'default_broker_id', context)
        profile = request.user.profile_user
        try:
            financial_broker = FinancialBroker.objects.get(id=default_broker_id, created_by=request.user)
            profile.user_financial_default_broker_id = default_broker_id
            profile.save()
            return JsonResponse({'message': 'done'})
        except:
            return JsonResponse({'message': 'broker not found or not belong to current user'})
    else:
        return JsonResponse({'message': 'not allowed'})
